"""
This is all about the automation of excel file with the help of python
"""
print("----------------------Automation of excel with python-------------------------")
from openpyxl import workbook, load_workbook

try:
    wb = load_workbook('example.xlsx')
except Exception as e:
    print(f"An error occurred: {e}")

ws = wb.active
index = input("Enter the index: ")
print(ws[index].value)


#for changing the value again to its original value
"""
ws['A1'].value = '4/5/2023'
wb.save('example.xlsx')
"""
print("------------------------------------------------------------------------------")
#this was for changing the value of Index C1 from 45 to 55
#ws[index].value = "55"


#this is for chaning the value after the save command
#wb.save('example.xlsx ')

#printing the changed value
#print(ws['c1'].value)
"""
wb = load_workbook('example.xlsx')
ws = wb.active"""

print("------------------------------------------------------------------------------")

"""print(wb['changable_sheet'])
ws['A1'].value = "Name"
wb.save('example.xlsx')
print(ws['A1'].value)"""
print("------------------------------------------------------------------------------")
